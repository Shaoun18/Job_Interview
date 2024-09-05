import selenium
import openpyxl
import datetime
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

# Get current day of the week
current_datetime = datetime.datetime.now()
current_day = current_datetime.strftime("%A")
print(f"Current Day: {current_day}")

# Load the Excel workbook and sheet corresponding to the current day
filename = r"4BeatsQ1.xlsx"
df = load_workbook(filename)
if current_day not in df.sheetnames:
    print(f"No sheet found for {current_day}")
    exit()

sheet = df[current_day]

# Number of keywords, or you can dynamically count the number of rows with keywords
num_keywords = 10  # Update this to the number of keywords in your sheet

# Set up ChromeDriver
service = Service(r'path_to_your_chromedriver')  # Update this path
driver = webdriver.Chrome(service=service)

# Iterate through the keywords
for row in range(3, 3 + num_keywords):
    keyword = sheet[f'C{row}'].value
    if not keyword:
        print(f"Empty keyword at row {row}, skipping.")
        continue

    print(f"Searching for keyword: {keyword}")

    # Perform Google search
    driver.get("https://www.google.com/")
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)
    search_box.submit()

    # Wait for search suggestions to load
    time.sleep(2)

    # Get suggestions
    options = driver.find_elements(By.XPATH, '//li[@role="presentation"]//span')

    if not options:
        print(f"No autocomplete options found for keyword: {keyword}")
        continue

    # Initialize longest and shortest with the first suggestion
    longest = options[0].text
    shortest = options[0].text
    len_longest = len(longest)
    len_shortest = len(shortest)

    # Find the longest and shortest suggestions
    for option in options[1:]:
        suggestion_text = option.text

        if len(suggestion_text) > len_longest:
            longest = suggestion_text
            len_longest = len(suggestion_text)

        if len(suggestion_text) < len_shortest:
            shortest = suggestion_text
            len_shortest = len(shortest)

    # Save the longest and shortest suggestions back to Excel
    sheet[f'D{row}'] = longest
    sheet[f'E{row}'] = shortest

    print(f"Longest search suggestion: {longest}")
    print(f"Shortest search suggestion: {shortest}")
    print()

# Close the browser and save the updated Excel file
driver.quit()
df.save(filename)
print("Keyword searching - Completed")