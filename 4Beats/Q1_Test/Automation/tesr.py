import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from datetime import datetime


def find_longest_shortest_options(driver, keyword):
    try:
        # Find the search box and input the keyword
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)  # Use RETURN to submit the search

        driver.implicitly_wait(2)  # Wait for suggestions to appear

        # Get the autocomplete suggestions
        suggestions = driver.find_elements(By.XPATH, '//li[@role="presentation"]//span')

        if not suggestions:
            print(f"No suggestions found for keyword: {keyword}")
            return None, None

        options = [suggestion.text for suggestion in suggestions]
        longest_option = max(options, key=len)
        shortest_option = min(options, key=len)
        return longest_option, shortest_option

    except Exception as e:
        print(f"Error during search for keyword '{keyword}': {e}")
        return None, None


def main():
    # Path to your Excel file
    excel_file_path = r'E:\Python\Interview4BeatsQ1.xlsx'

    # Load the Excel workbook
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # Determine the current day of the week
    day_of_week = datetime.now().strftime('%A')
    try:
        sheet = workbook[day_of_week]  # Access the sheet for the current day
    except KeyError:
        print(f"No sheet found for {day_of_week}")
        return

    # Set up the WebDriver
    try:
        service = Service(r'C:\Users\User\.wdm\drivers\chromedriver\win64\128.0.6613.86\chromedriver-win32')
        driver = webdriver.Chrome(service=service)
        driver.get("https://www.google.com")
    except Exception as e:
        print(f"Error setting up WebDriver: {e}")
        return

    # Iterate through keywords and perform the search
    for row in sheet.iter_rows(min_row=2, max_col=1):
        keyword = row[0].value
        if keyword:
            print(f"Searching for keyword: {keyword}")
            longest, shortest = find_longest_shortest_options(driver, keyword)
            if longest and shortest:
                sheet.cell(row=row[0].row, column=2).value = longest
                sheet.cell(row=row[0].row, column=3).value = shortest
                print(f"Longest: {longest}, Shortest: {shortest}")

    # Save the workbook
    try:
        workbook.save(excel_file_path)
        print(f"Workbook saved successfully at {excel_file_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

    # Close the browser
    driver.quit()


if __name__ == "__main__":
    main()
